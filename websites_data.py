import time
import argparse
import aiohttp
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from PIL import Image
from io import BytesIO
import base64
import requests
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# 初始化浏览器驱动
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # 隐藏浏览器窗口
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(options=options)
    return driver

# 异步获取状态码
async def get_status_code(url):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            return response.status

'''
# 设置 Chrome 选项
chrome_options = Options()
chrome_options.add_argument('--headless')  # 无头模式，不打开浏览器窗口
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# 初始化 WebDriver
def init_driver():
    service = Service('/usr/local/bin/chromedriver')  # 将 path_to_chromedriver 替换为你的 chromedriver 路径
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver
'''


# 访问单个网站，获取标题和可视区域截图
def fetch_site_data(driver, url):
    try:
        driver.get(url)
        time.sleep(3)  # 等待页面加载
        
        # 获取标题
        title = driver.title
        
        # 截取网页可视区域截图并转换为base64格式
        screenshot = driver.get_screenshot_as_base64()
        
        # 获取状态码
        status_code = requests.get(url).status_code
        
        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        return {"url": url, "title": title, "screenshot": screenshot, "status_code": status_code, "time": current_time}
    
    except TimeoutException:
        print(f"Timeout while loading {url}")
        return {"url": url, "title": "Timeout", "screenshot": "", "status_code": "Timeout", "time": ""}
    
    except WebDriverException as e:
        print(f"Error accessing {url}: {str(e)}")
        return {"url": url, "title": "Error", "screenshot": "", "status_code": "Error", "time": ""}

# 将数据和截图插入到 Excel，自动调整行高
def save_to_excel(data, file_name="websites_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Websites Data"

    # 设置列标题
    headers = ["序号", "url", "标题title", "状态码", "截图", "时间"]
    ws.append(headers)

    # 设置图片的理想宽度（高度自动按比例调整）
    ideal_image_width = 300  # 图片的理想宽度

    for idx, row_data in enumerate(data, 2):  # 从第二行开始插入数据
        ws[f"A{idx}"] = row_data["序号"]
        ws[f"B{idx}"] = row_data["url"]
        ws[f"C{idx}"] = row_data["标题title"]
        ws[f"D{idx}"] = row_data["状态码"]
        ws[f"F{idx}"] = row_data["时间"]

        # 将截图插入到Excel表中
        if row_data["screenshot"]:
            screenshot_data = base64.b64decode(row_data["screenshot"])
            screenshot_image = Image.open(BytesIO(screenshot_data))

            # 根据理想宽度调整图片尺寸
            img_width, img_height = screenshot_image.size
            scale_ratio = ideal_image_width / img_width  # 计算缩放比例
            img_resized_width = ideal_image_width
            img_resized_height = int(img_height * scale_ratio)  # 保持图片比例


            # 保存调整后的截图到内存
            img_stream = BytesIO()
            screenshot_image = screenshot_image.resize((img_resized_width, img_resized_height), Image.LANCZOS)
            screenshot_image.save(img_stream, format='PNG')
            img_stream.seek(0)

            # 插入图片
            img = ExcelImage(img_stream)
            img.width, img.height = img_resized_width, img_resized_height  # 设置图片大小
            ws.add_image(img, f"E{idx}")

            # 根据图片高度自动调整行高
            ws.row_dimensions[idx].height = img_resized_height * 0.75  # 适应Excel的行高比例

    # 调整列宽
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        if col == 5:  # 截图列
            ws.column_dimensions[col_letter].width = 50  # 宽度适合显示图片
        else:
            ws.column_dimensions[col_letter].width = 30  # 其他列的宽度

    wb.save(file_name)
    print(f"Data saved to {file_name}")

# 主函数，批量访问网站并保存数据
def main(urls_file):
    driver = init_driver()

    data_list = []
    
    # 读取URL文件中的所有网址
    with open(urls_file, 'r') as f:
        urls = f.readlines()

    total_urls = len([url for url in urls if url.strip()])  # 计算有效的URL数量

    # 遍历所有网址
    for idx, url in enumerate(urls, 1):
        url = url.strip()  # 去掉URL中的换行符和空格
        if not url:  # 跳过空行
            continue
        
        print(f"Processing {idx}/{total_urls}: {url}")  # 打印进度信息

        site_data = fetch_site_data(driver, url)
        
        # 保存数据（不保存截图文件到磁盘）
        data_list.append({
            "序号": idx,
            "url": site_data['url'],
            "标题title": site_data['title'],
            "状态码": site_data['status_code'],
            "screenshot": site_data['screenshot'],
            "时间": site_data['time']
        })
    
        # print(f"Completed {idx}/{total_urls}: {url}")  # 打印完成信息
    
    driver.quit()
    
    # 保存到 Excel
    save_to_excel(data_list)

if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='批量访问网站并保存信息到Excel')
    parser.add_argument('-u', '--urls', required=True, help='包含网站URL的文本文件路径')
    args = parser.parse_args()

    # 读取URL文件并执行主函数
    main(args.urls)